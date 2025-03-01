#   --New-- 
# altered attendance GUI: Added a text widget to the GUI to be able to display messages 
# write_nfc: Added display messages: Who's data is being added, and if it was successful
# read_nfc: Added display message to notify the user that theres no cin assigned which prompts them to assign accordingly
# log_attendance: Added display messages to confirm students attendance and if theyve already been logged
# imported messagebox from tkinter for error handling messages
# For all exceptions, created an error message box for the user to see
# __main__: display messages
# attenance gui: added search functionality, 
        # refractored functions for organization and to avoid confliting trees
        # added text widgets for pop ups
        # Added exit_gui function to ensure the program ends if the gui is gone
# Added message box to the start of the program to prompt the user 
# Added if in case they choose not to open a file which ends the program
# Added new global variables
        # previous_status : to prevent Waiting for Tag... message to keep repeating :b
        # cin_not_recorded_message_displayed : to prevent NFC does not have CIN# recorded repeatidly  

import csv                                                      # Import the CSV module for reading and writing CSV files
import os                                                       # Import the OS module for file and path operations
import time                                                     # Import the time module for adding delays
from datetime import datetime                                   # Import datetime for timestamp creation
from smartcard.System import readers                            # Import readers to interact with NFC readers
from smartcard.util import toHexString                          # Import toHexString to convert byte data to hex string
from smartcard.Exceptions import NoCardException                # Import NoCardException for error handling
import tkinter as tk                                            # Import tkinter for creating the GUI
#import tkinterFileDialog                                       # Import the FileDialog for User to pick file path for csv_path and folder_path
from tkinter import ttk, messagebox, filedialog, simpledialog   # Import ttk for themed tkinter widgets
import threading                                                # Import threading for running NFC reading in a separate thread
import openpyxl
from smartcard.util import toBytes
# from flask import Flask                             # Import flask for web application

# Show message box to prompt user for master list file
messagebox.showinfo("Select Master List", "Please select the Excel file containing the master list of students.")

# student_list = tk.filedialog.askopenfilename(title="yourmom", initaldir=os.path.expanduser('~'))   
# print(student_list)

# Now open the file dialog
file_path = filedialog.askopenfilename(
        title="Master List of Students",
        filetypes=[("Excel Files", "*.xlsx"), ("All files", "*.*")],
    )
if not file_path:
    messagebox.showerror("Error", "No file selected. The program will now exit.")
    exit()
print(file_path)

# eventName = input('Enter a name of the event (NO SPACES): ')
# eventName = "WelcomeMixer"
# Function to get event name and folder location
def get_event_details():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    event_name = simpledialog.askstring("Event Name", "Enter the name of the event (No Spaces):")
    if event_name is None:  # User cancelled
        exit()
    folder_path = filedialog.askdirectory(title="Select folder for event data")
    if not folder_path:  # User cancelled
        exit()
    return event_name, folder_path

# def get_student_list():
#     root = tk.Tk()
#     root.withdraw()
#     student_list = filedialog.askopenfilename("")

# Get event details
eventName, folder_path = get_event_details()

# Get one drive path details
# fileName, oneDrive_folderPath = get_student_list()

# Define custom event
CUSTOM_EVENT = '<<AttendanceLogged>>'

# Set the path for the attendance CSV file
# csv_path = os.path.join(os.path.expanduser('~'), 'Documents', 'Fall2024Events', eventName, 'attendance.csv') 
csv_path = os.path.join(folder_path, f"{eventName}_attendance.csv")

# Path to Excel file in OneDrive Folder for Students
# onedrive_path = os.path.join(os.path.expanduser('~'), 'OneDrive - Cal State LA', 'Registered ECST Transfers.xlsx')
onedrive_path = file_path

# Creating the flask app and URL's path component at root for the app
# app = Flask(__name__)
# @app.route("/")

# The line creates a path that points to a file named attendance.csv in the Documents folder of the user's home directory.
# os.path.expanduser('~'):
#     ~ is a shorthand in many operating systems for the user's home directory.
#     os.path.expanduser() expands this ~ to the full path of the user's home directory.
#     For example, on Windows, it might expand to C:\Users\YourUsername, and on macOS or Linux, it might be /home/YourUsername.
# 'Documents':
#     This is a standard folder name found in most user home directories.
# 'attendance.csv':
#     This is the name of the CSV file where attendance data will be stored.
# os.path.join():
#     This function is used to join multiple path components intelligently.
#     It handles the correct use of path separators (/ or $$ depending on the operating system.
# Advantages:
#     It's cross-platform compatible, working on Windows, macOS, and Linux.
#     It uses the user's home directory, so it doesn't require hard-coding a specific path.
#     It places the file in a standard location (Documents folder) where the user can easily find it.

# Global Variables
def globalVar():
    global stated
    stated = None

    global existing_entries
    existing_entries = []

    global signIn_statedAlready
    signIn_statedAlready = False

    global readerStatusStated
    readerStatusStated = None

    global studentData
    studentData = ""

    global previous_status
    previous_status = None

    global display_noCin
    display_noCin = True

# Function to get data from registered students excel
def get_registered_student_from_excel(rowNumber):
    print(f"Attempting to read from Excel file: {onedrive_path}")
    try:
        workbook = openpyxl.load_workbook(onedrive_path)
        sheet = workbook.active
        max_row = sheet.max_row
        print(f"Successfully opened workbook. Sheet has {sheet.max_row} rows.")

        rowNumber = int(rowNumber)
        if rowNumber < 2 or rowNumber > max_row:  # Check if row is in valid range
            raise ValueError("Row number out of range")
        firstName = sheet.cell(row=rowNumber, column=1).value
        lastName = sheet.cell(row=rowNumber, column=2).value
        cin = sheet.cell(row=rowNumber, column=3).value
        major = sheet.cell(row=rowNumber, column=4).value
        if not all([firstName, lastName, cin, major]):  # Check if any field is empty
            raise ValueError("Incomplete data in row")
        
        return cin, firstName, lastName, major
    
    except ValueError as e:
        messagebox.showerror("Error", "{e}")
        print(f"Error: {e}")
        return None, None, None, None

    except FileNotFoundError:
            messagebox.showerror("Error", "File not found. Retrying in 5 seconds...")
            print(f"File not found. Retrying in 5 seconds...")
            time.sleep(5)

    except Exception as e:  # Handle any other exceptions
        messagebox.showerror("Error", "File not found. Retrying in 5 seconds...")
        print(f"Error reading card: {e}")

# Function to write NFC tag
def write_nfc(firstName, lastName, cin, major):
    global readerStatusStated
    
    connection = connectReader()

    if not connection:
        print("Failed to connect to reader")
        return False

    try:
        connection.connect()
    except Exception as e:
        app.display_message(f"Failed to connect to card: {e}")
        print(f"Failed to connect to card: {e}")
        return False
    
    # Allows the user to see that theyre assigned the NFC Tag
    app.display_message(f"Writing data for {firstName} {lastName}...")

    # Format the data to match the expected read format
    data = f"CinNumber{cin}FirstName{firstName}LastName{lastName}Major{major}End"
    # print(f"Data to write: {data}")
    start_block = 4
    max_length = 121

    if len(data) > max_length:
        print(f"Data too long ({len(data)} bytes). Maximum is {max_length} bytes.")
        return False

    # Convert string to bytes
    data_bytes = data.encode('ascii')
    
    # Write data in 4-byte blocks
    for i in range(0, len(data_bytes), 4):
        block = start_block + (i // 4)
        chunk = data_bytes[i:i+4].ljust(4, b'\x00')  # Pad with null bytes if needed

        write_command = [0xFF, 0xD6, 0x00, block, 0x04] + list(chunk)
        # print(f"Writing to block {block}: {write_command}")

        try:
            response = connection.transmit(write_command)
            if response[1] != 0x90 or response[2] != 0x00:
                print(f"Write failed for block {block}. Response: {response}")
                return False
        except Exception as e:
            app.show_error("Error", f"Error writing to block {block}: {e}")
            print(f"Error writing to block {block}: {e}")
            return False

    print("Write operation completed successfully")
    app.display_message("Write operation completed successfully")

    return True

# Function to establish connection
def connectReader():
    global readerStatusStated

    r = readers()  # Get a list of available NFC readers
    if len(r) < 1:  # Check if any readers are available
        print("No reader found")
        app.display_message("No Reader bruh")
        readerStatusStated = None
        return None

    reader = r[0]  # Select the first available reader
    if readerStatusStated == None:
        print(f"Using reader: {reader}")
        app.display_message(f"Using reader: {reader}")
        readerStatusStated = True

    try:
        connection = reader.createConnection()  # Create a connection to the reader
    except Exception as e:
        app.display_message(f"Unexpected error: {str(e)}")
    
    return connection

def process_row_input(row_number):
    try:
        cin, firstName, lastName, major = get_registered_student_from_excel(row_number)
        print(f"Writing data for {firstName} {lastName}...")

        connection = connectReader()
        if not connection:
            print("Failed to connect to reader")
            return
        if write_nfc(firstName, lastName, cin, major):
            print("Data written successfully. Please tap the NFC tag again to log attendance.")
        else:
            print("Failed to write data to NFC tag.")
    except Exception as e:
        print(f"Unexpected error: {str(e)}")

# Function to read NFC tag
def read_nfc():
    global stated, readerStatusStated, existing_entries, signIn_statedAlready, display_noCin
    
    connection = connectReader()

    # Prevents from repeatedly saying that the card is not detected
    if stated != False and stated != True: 
        stated = False

    try:
        connection.connect()  # Connect to the NFC card

        # First, try to get the UI
        # Applying this APDU command to this 'variable'
        get_uid = [0xFF, 0xCA, 0x00, 0x00, 0x00] 
        
        # Basically, transmits the APDU command above onto the reader.
        # The result we get is FIXED, documentation pg 11
        # We retrieve the data, sw1 and sw2. IN THAT ORDER. That's how it assigns perfectly.
        data, sw1, sw2 = connection.transmit(get_uid)
        
        # # If the data in sw1 is 90, then operation sucess.
        # if sw1 == 0x90:
        #     uid = toHexString(data).replace(" ", "")
        #     print(f"UID: {uid}")
        # else:
        #     print(f"Error getting UID: {hex(sw1)}, {hex(sw2)}")
         
        #GetData
        #Read in sections of 4 bytes starting from block 4
        result = ""
        for block in range(4, 50):  # Read blocks 4 to 49
            read_data = [0xFF, 0xB0, 0x00, block, 0x04]  # Read 4 bytes from each block
            data, sw1, sw2 = connection.transmit(read_data)
            # print(f"Block: {block} Data: {data}")
            if sw1 == 0x90:
                # Convert data to ASCII, ignore non-printable characters
                ascii_data = ''.join([chr(b) for b in data if 32 <= b <= 126])
                # print (f"Block: {block} = {ascii_data}") # This prints what each block byte holds (16bytes each block)
                result += ascii_data # stores entire data into a string
            else:
                # print(f"Error reading block {block}: {hex(sw1)}, {hex(sw2)}")
                return None, None, None, None, None
        # print(f"Raw Data: {result}")
        
        if sw1 == 0x90:
            # Find the start of the CinNumber data
            cin_start = result.find("CinNumber")
            if cin_start != -1:
                # Extract only the numeric part after "CinNumber"
                cin_number = ''.join(filter(str.isdigit, result[cin_start+9:]))
                if stated == True:
                    stated = False
            else:
                if display_noCin:
                    app.display_message("NFC does not have a CIN# recorded in the data. \nPlease input a row number to assign data.")
                    print("NFC does not have a CIN# recorded in the data")
                    display_noCin = False
                return ("EMPTY", None, None, None, None)

            firstName_start = result.find("FirstName")
            lastName_start = result.find("LastName")
            major_start = result.find("Major")
            end = result.find("End")

            fN_length = lastName_start - (firstName_start+9)
            lN_length = major_start - (lastName_start+8)
            maj_length = end - (major_start+5)

            if firstName_start != -1:
                # Extract only the numeric part after "FirstName"
                firstName = result[firstName_start+9:firstName_start+9+fN_length]
                
                if stated == True:
                    stated = False
            else:
                print("FirstName not found in data")

            if lastName_start != -1:
                # Extract only the numeric part after "CinNumber"
                lastName = result[lastName_start+8:lastName_start+8+lN_length]

                if stated == True:
                    stated = False
            else:
                print("LastName not found in data")

            major_start = result.find("Major")
            if major_start != -1:
                # Extract only the numeric part after "CinNumber"
                major = result[major_start+5:major_start+5+maj_length]
                
                if stated == True:
                    stated = False
            else:
                print("Major not found in data")     
        else:
            print(f"Error reading data: {hex(sw1)}, {hex(sw2)}")

        if is_cin_recorded(cin_number):
            if not signIn_statedAlready: 
                print(f"{firstName} {lastName} has already signed in. \n")
                app.display_message(f"\n{firstName} {lastName} has already signed in. \n")
                signIn_statedAlready = True
        else:
            signIn_statedAlready = False

            # print(f"\nCIN#: {cin_number}")
            # print(f"First Name: {firstName}")
            # print(f"Last Name: {lastName}")                
            # print(f"Major: {major}")    

        if cin_number and firstName and lastName and major:
            signIn_statedAlready = True
            return ("SUCCESS", cin_number, firstName, lastName, major)
                
    except NoCardException:  # Handle the case when no card is detected
        if stated == False:
            app.display_message("No card detected. \nPlease place a card on the reader.")
            print("No card detected. Please place a card on the reader.")
            stated = True
            display_noCin = True
        return ("NO_CARD", None, None, None, None)
        
    except Exception as e:  # Handle any other exceptions
        # print(f"Error reading card: {e}")
        app.display_message("Card has been removed.")
        print("Card has been removed.")
        
        if signIn_statedAlready:
                signIn_statedAlready = False
        return ("ERROR", None, None, None, None)
    
# Function to log attendance
def log_attendance(student_cin, student_firstName, student_lastName, student_major):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Get current timestamp
    global existing_entries

    # If CIN doesn't exist, add new entry
    if not is_cin_recorded(student_cin):
        # Log new attendance        
        existing_entries.append(student_cin) # Add student CIN into global array
        with open(csv_path, 'a', newline='') as file:  # Open the CSV file in append mode
            writer = csv.writer(file)  # Create a CSV writer object
            writer.writerow([student_cin, student_firstName, student_lastName, student_major, timestamp])  # Write the attendance record
        app.display_message(f"\nLogged attendance for {student_firstName} {student_lastName} at {timestamp}\n")
        print(f"Logged attendance for {student_firstName} {student_lastName} at {timestamp}\n")
    else:
        app.display_message(f"CIN {student_cin} already recorded.")
        print(f"CIN {student_cin} already recorded.") 
        for entry in existing_entries:
            print(entry)
    return student_cin, student_firstName, student_lastName, student_major, timestamp  # Return the logged data

def is_cin_recorded(cin):
    return cin in existing_entries

def initialize_csv(root):
    global studentData
    
    # Ensure the directory exists
    directory = os.path.dirname(csv_path)   # Checks if the correct path exists
    try:
        if not os.path.exists(directory):       # If not, the directory will be made
            print("No directory found.")
            os.makedirs(directory)
            print(f"Created directory: {directory}")
    except OSError as e:
        messagebox.showerror("Error", "Error creating a directory")
        print("Error creating a directory/n")
    
    # Check if the CSV file already exists
    if not os.path.exists(csv_path):                        
        with open(csv_path, 'w', newline='') as file:       # If not, create a new CSV file
            writer = csv.writer(file)                       # Create a CSV writer object
            writer.writerow(["Student CIN", "First Name", "Last Name", "Major" ,"Timestamp"])    # Write the header row
        print("Created new attendance CSV file.")
    else:
        print("Attendance CSV file already exists.")
        #Already recorded entries of the excel file is added into global array
        with open(csv_path, 'r', newline='') as file:
            reader = csv.reader(file)
            next(reader)  # Skip header
            for row in reader:
                existing_entries.append(row[0])
                studentData = f"{row[0]},{row[1]},{row[2]},{row[3]},{row[4]}"
                root.event_generate(CUSTOM_EVENT, when='now')

def load_excel_data(gui):
    try:
        workbook = openpyxl.load_workbook(onedrive_path)
        sheet = workbook.active
        for row in range(2, sheet.max_row + 1):
            firstName = sheet.cell(row=row, column=1).value
            lastName = sheet.cell(row=row, column=2).value
            cin = sheet.cell(row=row, column=3).value
            major = sheet.cell(row=row, column=4).value
            gui.excel_tree.insert('', 'end', values=(row, firstName, lastName, cin, major))
    except Exception as e:
        app.show_error("Error", f"Error loading Excel data: {e}")
        print(f"Error loading Excel data: {e}")

class AttendanceGUI:
    def show_error(self, title, message):
        self.master.after(0, lambda: messagebox.showerror(title, message))

    def create_excel_treeview(self):
        self.excel_tree = ttk.Treeview(self.excel_frame, columns=('Row', 'First Name', 'Last Name', 'CIN', 'Major'), show='headings')
        for col in ['Row', 'First Name', 'Last Name', 'CIN', 'Major']:
            self.excel_tree.heading(col, text=col)
        self.excel_tree.pack(fill=tk.BOTH, expand=1)

    def create_search_functionality(self):
        self.search_frame = ttk.Frame(self.excel_frame)
        self.search_frame.pack(fill=tk.X, padx=10, pady=10)
        self.search_label = ttk.Label(self.search_frame, text="Search CIN:")
        self.search_label.pack(side=tk.LEFT)
        self.search_entry = ttk.Entry(self.search_frame)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        self.search_button = ttk.Button(self.search_frame, text="Search", command=self.search_cin)
        self.search_button.pack(side=tk.LEFT)
    
    def __init__(self, master):
        self.master = master
        master.title("Attendance Tracker")

        # Create notebook for tabs
        self.notebook = ttk.Notebook(master)
        self.notebook.pack(fill=tk.BOTH, expand=1)

        # Attendance Tab
        self.attendance_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.attendance_frame, text='Attendance')

        self.tree = ttk.Treeview(self.attendance_frame, columns=('Student CIN', 'First Name', 'Last Name', 'Major', 'Timestamp'), show='headings')
        self.tree.heading('Student CIN', text='Student CIN#')
        self.tree.heading('First Name', text='First Name')
        self.tree.heading('Last Name', text='Last Name')
        self.tree.heading('Major', text='Major')
        self.tree.heading('Timestamp', text='Timestamp')
        self.tree.pack(fill=tk.BOTH, expand=1)

        # Excel Data Tab
        self.excel_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.excel_frame, text='Excel Data')
        
        self.create_excel_treeview()
        self.create_search_functionality()

        # Refresh button for Excel data
        self.refresh_button = ttk.Button(self.excel_frame, text="Refresh Excel Data", command=self.refresh_excel_data)
        self.refresh_button.pack()

        # Input frame for row number
        self.input_frame = ttk.Frame(master)
        self.input_frame.pack(fill=tk.X, padx=10, pady=10)

        self.row_label = ttk.Label(self.input_frame, text="Enter Row Number:")
        self.row_label.pack(side=tk.LEFT)

        self.row_entry = ttk.Entry(self.input_frame)
        self.row_entry.pack(side=tk.LEFT, padx=5)

        self.submit_button = ttk.Button(self.input_frame, text="Submit", command=self.submit_row)
        self.submit_button.pack(side=tk.LEFT)

        self.master.bind(CUSTOM_EVENT, self.handle_attendance_logged)

        # Add a text widget for messages
        self.message_text = tk.Text(master, height=10, width=50)
        self.message_text.pack(pady=10)

        # tk toolkit, window close event
        self.master.protocol("WM_DELETE_WINDOW",self.close_gui)

    # "destroys" the gui
    def close_gui(self):
        self.display_message("\nExiting Program...\n")
        time.sleep(2)
        self.master.destroy()
        raise SystemExit("GUI exited")

    def display_message(self, message):
        self.message_text.insert(tk.END, message + "\n")
        self.message_text.see(tk.END)  # Scroll to the end

    def handle_attendance_logged(self, event):
        try:
            student_id, student_firstName, student_lastName, student_major, timestamp = studentData.split(',')
            self.tree.insert('', 'end', values=(student_id, student_firstName, student_lastName, student_major, timestamp))
        except Exception as e:
            app.show_error("Error", f"Error handling attendance event: {e}")
            print(f"Error handling attendance event: {e}")

    def refresh_excel_data(self):
        self.excel_tree.delete(*self.excel_tree.get_children())
        load_excel_data(self)

    def search_cin(self):
        search_cin = self.search_entry.get()
        if search_cin:
            for item in self.excel_tree.get_children():
                values = self.excel_tree.item(item)['values']
                if str(values[3]) == search_cin:  # CIN is at index 3
                    self.excel_tree.selection_set(item)
                    self.excel_tree.focus(item)
                    self.excel_tree.see(item)
                    return
            messagebox.showinfo("Search Result", f"No student found with CIN: {search_cin}")
        else:
            messagebox.showwarning("Search Error", "Please enter a CIN to search")

    def submit_row(self):
        row_number = self.row_entry.get()
        if row_number:
            process_row_input(row_number)
        self.row_entry.delete(0, tk.END)

def main_loop():
    app.display_message("Reached main loop")
    global studentData, previous_status
    try:
        while True:
            status, cin, firstName, lastName, major = read_nfc()
            display = False

            if status != previous_status:
                if status == "NO_CARD":
                    app.display_message("Waiting for a tag...")
                    print("Waiting for a tag...", end='\r', flush=True)
                previous_status = status

            if status == "EMPTY":
                root.event_generate('<<EmptyNFC>>')  # Generate custom event for empty NFC
                app.master.after(0, lambda: root.event_generate('<<EmptyNFC>>'))
            elif status == "SUCCESS":
                if not is_cin_recorded(cin):
                    display = True
                    student_id, student_firstName, student_lastName, student_major, timestamp = log_attendance(cin, firstName, lastName, major)
                    if display:
                        studentData = f"{student_id},{student_firstName},{student_lastName},{student_major},{timestamp}"
                        root.event_generate(CUSTOM_EVENT, when='now')
            time.sleep(0.5)
    except SystemExit as e:
        print(f"\n{e}")
        app.display_message(f"\n{e}")
    except KeyboardInterrupt:
        print("\nScript stopped by user.")
        app.display_message("\nScript stopped by user.")

if __name__ == "__main__":
    globalVar()
    root = tk.Tk()
    app = AttendanceGUI(root)
    initialize_csv(root)
    load_excel_data(app)  # Load initial Excel data

    app.display_message("NFC Reader initialized.")
    app.display_message(f"Attendance is being logged to: {csv_path}\n")
    print("NFC Reader initialized.")
    print(f"Attendance is being logged to: {csv_path}")
    print("Press Ctrl+C in the console to stop the script.")

    app.display_message("Attempting Threading")
    nfc_thread = threading.Thread(target=main_loop, daemon=True)
    
    nfc_thread.start()
    app.display_message("Threading successful")

    try:
        root.mainloop()
    except SystemExit:
        # Goes to this if the GUI was closed which should make the exit safe and succesful
        pass  

    # Wait for the NFC thread to finish, about 2 seconds :)
    nfc_thread.join(timeout=2)

    print("\n***Program exited.***\n")

    # Use in terminal to create executable in the terminal of the folder (In the smae location where the python application is at)
    # pyinstaller --onefile --windowed --hidden-import smartcard --hidden-import openpyxl --hidden-import smartcard.System --hidden-import smartcard.util --hidden-import tkinter --hidden-import threading PythonApplication.py